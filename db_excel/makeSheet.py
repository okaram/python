
from openpyxl import Workbook

from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Style, PatternFill, Font

wb = Workbook()

ws = wb.create_sheet()

ws.title = 'Pi'

ws['F5'] = 3.14
ws['F5'].style = Style(fill=PatternFill(patternType='solid', fgColor=Color('AAFF0000')))


wb.save(filename = "test1.xlsx")

