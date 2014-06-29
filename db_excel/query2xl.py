import psycopg2 as dbapi2
import sys

from openpyxl import Workbook
#from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Style, PatternFill

headerStyle=Style(fill=PatternFill(patternType='solid', fgColor=Color('008080FF')))

class QuerySpec(object):
    def __init__(self, title, db, query):
        self.title=title
        self.db=db
        self.query=query
    
class SheetSpec(object):
    def __init__(self, title, queries):
        self.title=title
        self.queries=queries
    def addToWorkbook(self, wkBook):
        theRow=0
        ws=wkBook.create_sheet()
        ws.title=self.title
        for q,qspec in enumerate(self.queries):
            cur = qspec.db.cursor()
            cur.execute (qspec.query)
            theRow=self.cursor2xl(qspec.title, cur,ws, theRow)
            cur.close()


    def cursor2xl(self, title,cur, sheet,firstRow):
        theRow=firstRow+1;
        sheet.cell(row=theRow,column=1).value=title
        theRow=theRow+1
        for f,field in enumerate(cur.description):
            sheet.cell(row=theRow,column=f+1).value=field.name
            sheet.cell(row=theRow,column=f+1).style=headerStyle
        for r,row in enumerate(cur.fetchall()):
            theRow=theRow+1
            for c,field in enumerate(row):
                sheet.cell(row=theRow,column=c+1).value=field
        return theRow+2;
    

class WorkBookSpec(object):
    def __init__(self, sheets):
        self.sheets=sheets
    def mkSpreadSheet(self, ):
        pass
    

def cursor2print(cur):
    for f,field in enumerate(cur.description):
        print f,field.name
    for r, row in enumerate(cur.fetchall()):
        for c,val in enumerate(row):
            print c,val    

def cursor2xl(cur, sheet):
    for f,field in enumerate(cur.description):
        sheet.cell(row=1,column=f+1).value=field.name
        sheet.cell(row=1,column=f+1).style=headerStyle
    for r,row in enumerate(cur.fetchall()):
        for c,field in enumerate(row):
            sheet.cell(row=r+2,column=c+1).value=field
        
        #sheet.cell(row=1,column=f+1).style.fill.start_color.index = "FF124191"
        #print sheet.cell(row=1,column=f+1).style
    
def query2xl(db,query,wkBook):
    cur = db.cursor()
    cur.execute (query);
    ws=wkBook.create_sheet()
    ws.title='hope this works'
    cursor2xl(cur,ws)

def query2print(db,query):
    cur = db.cursor()
    cur.execute (query);
    cursor2print(cur)
    
if __name__ == '__main__':
    db = dbapi2.connect (database="curri", user="curri")
    wb = Workbook()
    query2xl(db,sys.argv[1],wb)
    sheet=SheetSpec('Products and People',
                    [QuerySpec('People',db,'Select * FROM Person'),
                    QuerySpec('Products',db,'Select * FROM Products')
                                           
                    ]
                    )
    sheet.addToWorkbook(wb);


    wb.save(filename = "test1.xlsx")
    