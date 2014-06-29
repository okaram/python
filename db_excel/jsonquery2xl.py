#import psycopg2 
import sys
import json

from openpyxl import Workbook
#from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, Style, PatternFill

headerStyle=Style(fill=PatternFill(patternType='solid', fgColor=Color('008080FF')))
    
def query2xl(query, sheet, firstRow, headerStyle):
    driver=__import__(query['db_driver'])
    db=driver.connect(query['db_conn_str'])
    cur=db.cursor()
    cur.execute(query['query'])
    theRow=firstRow+1;
    sheet.cell(row=theRow,column=1).value=query['title']
    #theRow=theRow+1
    for f,field in enumerate(cur.description):
        sheet.cell(row=theRow,column=f+2).value=field.name
        sheet.cell(row=theRow,column=f+2).style=headerStyle
    for r,row in enumerate(cur.fetchall()):
        theRow=theRow+1
        for c,field in enumerate(row):
            sheet.cell(row=theRow,column=c+2).value=field
    cur.close()
    db.close()
    return theRow+2;

def addSheetToWorkbook(sheet, wkBook):
    theRow=0
    ws=wkBook.create_sheet()
    ws.title=sheet['title']
    for q,query in enumerate(sheet['queries']):
        theRow=query2xl(query, ws,theRow,headerStyle)
    
def mkWorkBook(sheets):  
    wb = Workbook()
    for s,sheet in enumerate(sheets):
        addSheetToWorkbook(sheet,wb)
    return wb

if __name__ == '__main__':
    spec=json.loads(open(sys.argv[1],'r').read())
    wb=mkWorkBook(spec['sheets'])
    wb.save(filename = sys.argv[2])
    